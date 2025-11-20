from openpyxl import load_workbook
import requests
from utils.api_client import APIClient
from utils.auth import get_auth_token
from utils.config import tenantId

def prepare_template():
    """Download template and copy data from sample file"""

    # Step 1: Get download URL from API
    token = get_auth_token('user')
    client = APIClient(token=token)

    with open('output/ids.txt', 'r') as f:
        for line in f:
            if line.startswith('Generated FileStore ID:'):
                file_store_id = line.split(':')[1].strip()
                break

    url = f"/filestore/v1/files/url?tenantId={tenantId}&fileStoreIds={file_store_id}"
    response = client.get(url)

    download_url = response.json()['fileStoreIds'][0]['url']
    print(f"Download URL obtained: {download_url[:80]}...")

    # Step 2: Download the template from S3
    print("\nDownloading template from S3...")
    template_response = requests.get(download_url)

    with open('output/template_downloaded.xlsx', 'wb') as f:
        f.write(template_response.content)

    print(f"Template downloaded: {len(template_response.content)} bytes")

    # Step 3: Load both files
    print("\nLoading files...")
    template_wb = load_workbook('output/template_downloaded.xlsx')
    sample_wb = load_workbook('utils/sample_boundary.xlsx', data_only=True, read_only=True)

    template_ws = template_wb['Boundary Data']
    sample_ws = sample_wb['Boundary Data']

    # Show template headers
    print("\nTemplate Headers (keeping these):")
    for col_idx in range(1, 13):
        header = template_ws.cell(row=1, column=col_idx).value
        if header:
            print(f"  Col {col_idx}: {header}")

    # Step 4: Copy data from sample (rows 2 onwards) to template
    print("\nCopying data from sample file (rows 2 onwards)...")

    # Collect data rows from sample
    data_rows = []
    for row in sample_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            data_rows.append(row)

    print(f"Found {len(data_rows)} data rows in sample file")

    # Write data to template starting from row 2
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            template_ws.cell(row=row_idx, column=col_idx, value=value)

    # Step 5: Save the template
    print("\nSaving template with copied data...")
    template_wb.save('output/template_downloaded.xlsx')

    print("✓ Template prepared successfully")

    # Verify
    print("\nVerifying template content:")
    verify_wb = load_workbook('output/template_downloaded.xlsx')
    verify_ws = verify_wb['Boundary Data']

    print("\nHeaders (Row 1):")
    for col_idx in range(1, 8):
        header = verify_ws.cell(row=1, column=col_idx).value
        print(f"  Col {col_idx}: {header}")

    print("\nFirst 3 data rows:")
    for row_idx in range(2, 5):
        row_data = []
        for col_idx in range(1, 8):
            value = verify_ws.cell(row=row_idx, column=col_idx).value
            if value:
                row_data.append(f"Col{col_idx}:{value}")
        print(f"  Row {row_idx}: {' | '.join(row_data)}")

if __name__ == "__main__":
    prepare_template()
