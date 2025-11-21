from openpyxl import load_workbook

def copy_data_from_sample():
    """Copy data from sample_boundary.xlsx (row 2 onwards) to template_downloaded.xlsx"""

    print("Loading sample_boundary.xlsx...")
    sample_wb = load_workbook("/home/shreya-kumar/API_Automation/output/sample_boundary.xlsx", data_only=True, read_only=True)

    print("Loading template_downloaded.xlsx...")
    template_wb = load_workbook("/home/shreya-kumar/API_Automation/output/template_downloaded.xlsx")

    sample_ws = sample_wb['Boundary Data']
    template_ws = template_wb['Boundary Data']

    print("Collecting data rows from sample...")

    # Collect all rows with data from sample (starting from row 2)
    data_rows = []
    for row in sample_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            data_rows.append(row)

    print(f"Found {len(data_rows)} non-empty data rows")

    print("Writing data to template...")
    # Write data to template starting from row 2
    for idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            template_ws.cell(row=idx, column=col_idx, value=value)

    print("Saving template...")
    template_wb.save("/home/shreya-kumar/API_Automation/output/template_downloaded.xlsx")

    print(f"Successfully copied {len(data_rows)} rows")

    # Verify
    print("\nVerifying updated template...")
    verify_wb = load_workbook("/home/shreya-kumar/API_Automation/output/template_downloaded.xlsx")
    verify_ws = verify_wb['Boundary Data']

    non_empty = sum(1 for row in verify_ws.iter_rows(values_only=True) if any(cell for cell in row))
    print(f"Total non-empty rows in template: {non_empty}")

    # Show first few rows
    print("\nFirst 6 rows in updated template:")
    for idx, row in enumerate(verify_ws.iter_rows(values_only=True), 1):
        if idx <= 6:
            print(f"Row {idx}: {' | '.join([str(c)[:30] if c else '' for c in row[:7]])}")

if __name__ == "__main__":
    copy_data_from_sample()
