from utils.api_client import APIClient
from utils.auth import get_auth_token
from utils.config import tenantId, BASE_URL
import pytest
import requests
import os
from openpyxl import load_workbook


def prepare_template_for_upload(token):
    """Automatically prepare template by downloading and populating with sample data"""
    from utils.api_client import APIClient

    client = APIClient(token=token)

    # Read Generated FileStore ID from ids.txt
    file_store_id = None
    with open('output/ids.txt', 'r') as f:
        for line in f:
            if line.startswith('Generated FileStore ID:'):
                file_store_id = line.split(':')[1].strip()
                break

    if not file_store_id:
        raise Exception("Generated FileStore ID not found in output/ids.txt")

    # Get download URL from API
    url = f"/filestore/v1/files/url?tenantId={tenantId}&fileStoreIds={file_store_id}"
    response = client.get(url)

    if response.status_code != 200:
        raise Exception(f"Failed to get download URL: {response.text}")

    download_url = response.json()['fileStoreIds'][0]['url']

    # Download template from S3
    print(f"  Downloading template from S3...")
    template_response = requests.get(download_url)

    with open('output/template_downloaded.xlsx', 'wb') as f:
        f.write(template_response.content)

    # Load both files
    template_wb = load_workbook('output/template_downloaded.xlsx')
    sample_wb = load_workbook('utils/sample_boundary.xlsx', data_only=True, read_only=True)

    template_ws = template_wb['Boundary Data']
    sample_ws = sample_wb['Boundary Data']

    # Copy data from sample (rows 2 onwards only, skip headers)
    # This preserves the downloaded template's headers in row 1
    data_rows = []
    for row in sample_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            data_rows.append(row)

    print(f"  Copying {len(data_rows)} data rows from sample to template...")

    # Write data to template starting from row 2 (row 1 headers are preserved)
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            template_ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the populated template
    template_wb.save('output/sample_boundary.xlsx')
    print(f"  Template prepared successfully")


@pytest.mark.order(8)
def test_file_upload():
    """Test uploading a file"""
    token = get_auth_token("user")

    sample_file = "output/sample_boundary.xlsx"

    # Always prepare template to ensure it matches the current hierarchy
    print("\nPreparing template for upload...")
    try:
        prepare_template_for_upload(token)
    except Exception as e:
        pytest.skip(f"Could not prepare template: {e}")

    # Prepare multipart form data
    files = {
        'file': (os.path.basename(sample_file), open(sample_file, 'rb'),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    data = {
        'tenantId': tenantId,
        'module': 'HCM-ADMIN-CONSOLE'
    }

    # Make direct request
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.post(
        f"{BASE_URL}/filestore/v1/files",
        files=files,
        data=data,
        headers=headers,
        verify=False
    )

    files['file'][1].close()

    assert response.status_code in [200, 201], f"File upload failed: {response.text}"

    data = response.json()
    file_store_id = data["files"][0]["fileStoreId"]

    print(f"File uploaded successfully: {file_store_id}")

    # Update file store ID (overwrite existing)
    with open("output/ids.txt", "r") as f:
        lines = f.readlines()

    with open("output/ids.txt", "w") as f:
        for line in lines:
            if line.startswith("Uploaded FileStore ID:"):
                f.write(f"Uploaded FileStore ID: {file_store_id}\n")
            else:
                f.write(line)
        # Add if not found
        if not any(line.startswith("Uploaded FileStore ID:") for line in lines):
            f.write(f"Uploaded FileStore ID: {file_store_id}\n")
