from openpyxl import load_workbook

def show_utils_sample():
    """Display the sample_boundary.xlsx file from utils folder"""
    file_path = "/home/shreya-kumar/API_Automation/utils/sample_boundary.xlsx"

    wb = load_workbook(file_path, data_only=True, read_only=True)

    print(f"File: {file_path}")
    print(f"Sheets: {wb.sheetnames}\n")
    print("=" * 120)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\nSHEET: {sheet_name}")
        print("=" * 120)

        # Count non-empty rows
        non_empty_rows = sum(1 for row in ws.iter_rows(values_only=True) if any(cell for cell in row))
        print(f"Total Rows: {ws.max_row}, Non-empty Rows: {non_empty_rows}, Columns: {ws.max_column}\n")

        # Show all non-empty rows
        print("All non-empty data:\n")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(cell for cell in row):
                if row_idx == 1:
                    print(f"Row {row_idx} (HEADERS):")
                    for idx, cell in enumerate(row, 1):
                        if cell:
                            print(f"  Col {idx}: {cell}")
                    print("-" * 120)
                else:
                    print(f"\nRow {row_idx}:")
                    for idx, cell in enumerate(row, 1):
                        if cell:
                            print(f"  Col {idx}: {cell}")

if __name__ == "__main__":
    show_utils_sample()
