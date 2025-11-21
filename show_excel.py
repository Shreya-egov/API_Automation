from openpyxl import load_workbook

def show_excel_content(file_path):
    """Display all sheets in an Excel file"""
    try:
        # Load workbook without styles to avoid formatting errors
        wb = load_workbook(file_path, data_only=True, read_only=True)

        print(f"Excel file: {file_path}")
        print(f"Number of sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {wb.sheetnames}\n")
        print("=" * 100)

        # Display each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'=' * 100}")
            print(f"SHEET: {sheet_name}")
            print(f"{'=' * 100}\n")

            ws = wb[sheet_name]

            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            print(f"Rows: {max_row}, Columns: {max_col}\n")

            # Display the data
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    # Header row
                    print("Headers:", " | ".join([str(cell) if cell is not None else "" for cell in row]))
                    print("-" * 100)
                else:
                    print(f"Row {row_idx}: {' | '.join([str(cell) if cell is not None else '' for cell in row])}")

            print("\n")

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    show_excel_content("/home/shreya-kumar/API_Automation/output/sample_boundary.xlsx")
