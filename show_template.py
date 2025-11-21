from openpyxl import load_workbook

def show_template_content(file_path):
    """Display template file content"""
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)

        print(f"Template file: {file_path}")
        print(f"Sheets: {wb.sheetnames}\n")
        print("=" * 120)

        for sheet_name in wb.sheetnames:
            print(f"\nSHEET: {sheet_name}")
            print("=" * 120)

            ws = wb[sheet_name]

            # Count non-empty rows
            non_empty_rows = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell for cell in row):
                    non_empty_rows += 1

            print(f"\nTotal Rows: {ws.max_row}, Non-empty Rows: {non_empty_rows}, Columns: {ws.max_column}\n")

            # Show all non-empty rows
            print("All non-empty rows:\n")
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                # Check if row has any non-empty cell
                if any(cell for cell in row):
                    if row_idx == 1:
                        # Header row
                        print(f"Row {row_idx} (HEADERS):")
                        for idx, cell in enumerate(row, 1):
                            if cell:
                                print(f"  Column {idx}: {cell}")
                        print("-" * 120)
                    else:
                        print(f"\nRow {row_idx}:")
                        for idx, cell in enumerate(row, 1):
                            if cell:
                                print(f"  Column {idx}: {cell}")

    except Exception as e:
        print(f"Error reading template file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    show_template_content("/home/shreya-kumar/API_Automation/output/template_downloaded.xlsx")
